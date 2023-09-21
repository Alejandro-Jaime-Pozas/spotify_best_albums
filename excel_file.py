import openpyxl

def add_album_data_excel(list_data):

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Select the default active sheet
    sheet = workbook.active

    # Set column headers
    sheet['A1'] = 'spot_top_album_id'
    sheet['B1'] = 'spot_top_album_artist_name'
    sheet['C1'] = 'accl_album_artist_name'
    sheet['D1'] = 'top_album_artist_score'
    sheet['E1'] = 'spot_top_album_name'
    sheet['F1'] = 'accl_album_name'
    sheet['G1'] = 'top_album_score'
    sheet['H1'] = 'top_album_overall_score'
    sheet['I1'] = 'year'
    sheet['J1'] = 'album_found'

    # Write data to the worksheet
    for row_index, row_data in enumerate(list_data, start=2):
        sheet.cell(row=row_index, column=1, value=row_data['spot_top_album_id'])
        sheet.cell(row=row_index, column=2, value=row_data['spot_top_album_artist_name'])
        sheet.cell(row=row_index, column=3, value=row_data['accl_album_artist_name'])
        sheet.cell(row=row_index, column=4, value=row_data['top_album_artist_score'])
        sheet.cell(row=row_index, column=5, value=row_data['spot_top_album_name'])
        sheet.cell(row=row_index, column=6, value=row_data['accl_album_name'])
        sheet.cell(row=row_index, column=7, value=row_data['top_album_score'])
        sheet.cell(row=row_index, column=8, value=row_data['top_album_overall_score'])
        sheet.cell(row=row_index, column=9, value=row_data['year'])
        sheet.cell(row=row_index, column=10, value=row_data['album_found'])

    # Save the Excel file
    workbook.save('spotify_album_data.xlsx')
    # Close the workbook
    workbook.close()

    print("Excel file 'spotify_album_data' created successfully.")